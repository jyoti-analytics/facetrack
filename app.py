import tkinter as tk
from tkinter import ttk
import cv2
import numpy as np
import pickle
import json
import os
import csv
from datetime import datetime
from PIL import Image, ImageTk

# Load models
with open("model/pca.pkl", "rb") as f:
    pca = pickle.load(f)
with open("model/svm.pkl", "rb") as f:
    model = pickle.load(f)
with open("model/student_info.json", "r") as f:
    student_info = json.load(f)

student_info = {int(k): v for k, v in student_info.items()}

student_details = {
    1: {"name": "Jyoti", "roll": "22501102", "dept": "Computer Science"},
    2: {"name": "Niharika", "roll": "22501103", "dept": "Computer Science"},
}

face_cascade = cv2.CascadeClassifier(
    cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

attendance_marked = set()

def mark_attendance(student_id, student_name):
    if student_id not in attendance_marked:
        attendance_marked.add(student_id)
        now = datetime.now()
        date = now.strftime("%Y-%m-%d")
        time = now.strftime("%H:%M:%S")
        os.makedirs("attendance", exist_ok=True)
        details = student_details.get(student_id, {})
        roll = details.get("roll", "")
        dept = details.get("dept", "")

        # ---- Save to CSV ----
        csv_file = f"attendance/attendance_{date}.csv"
        file_exists = os.path.isfile(csv_file)
        with open(csv_file, "a", newline="") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(["ID","Name","Roll","Department","Date","Time","Status"])
            writer.writerow([student_id, student_name, roll, dept, date, time, "Present"])

        # ---- Save to Excel ----
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        excel_file = f"attendance/attendance_{date}.xlsx"

        if os.path.isfile(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Attendance {date}"

            # Header style
            headers = ["ID","Name","Roll No","Department","Date","Time","Status"]
            header_fill = PatternFill(start_color="FF4757",
                                      end_color="FF4757",
                                      fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Column widths
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 12

        # Add data row
        row = ws.max_row + 1
        data = [student_id, student_name, roll, dept, date, time, "Present"]
        green_fill = PatternFill(start_color="2ED573",
                                  end_color="2ED573",
                                  fill_type="solid")
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")
            if col == 7:
                cell.fill = green_fill
                cell.font = Font(bold=True)

        wb.save(excel_file)
        return True
    return False
class FaceTrackApp:
    def __init__(self, root):
        self.root = root
        self.root.title("FACETRACK - Intelligent Attendance System")
        self.root.configure(bg="#0d0d0d")
        self.root.state('zoomed')
        self.cap = None
        self.running = False
        self.build_ui()

    def build_ui(self):
        # ---- TOP BAR ----
        top = tk.Frame(self.root, bg="#111111", height=50)
        top.pack(fill="x", side="top")
        top.pack_propagate(False)

        tk.Label(top, text="FACETRACK — Intelligent Real-Time Attendance System",
                 font=("Arial", 15, "bold"),
                 bg="#111111", fg="#ff4757").pack(side="left", padx=20, pady=10)

        self.datetime_label = tk.Label(top, text="",
                                       font=("Arial", 11),
                                       bg="#111111", fg="#aaaaaa")
        self.datetime_label.pack(side="right", padx=20)
        self.update_datetime()

        # ---- BOTTOM BUTTONS BAR ----
        bottom = tk.Frame(self.root, bg="#111111", height=65)
        bottom.pack(fill="x", side="bottom")
        bottom.pack_propagate(False)

        self.status_label = tk.Label(bottom, text="● Ready — Press START to begin",
                                     font=("Arial", 10, "bold"),
                                     bg="#111111", fg="#2ed573")
        self.status_label.pack(side="left", padx=20, pady=10)

        # Buttons on right side of bottom bar
        btn_frame = tk.Frame(bottom, bg="#111111")
        btn_frame.pack(side="right", padx=20, pady=10)

        tk.Button(btn_frame, text="✕  CLEAR LOG",
                  font=("Arial", 11, "bold"),
                  bg="#222222", fg="#aaaaaa",
                  padx=16, pady=6, relief="flat",
                  cursor="hand2",
                  command=self.clear_log).pack(side="right", padx=5)

        self.stop_btn = tk.Button(btn_frame, text="■  STOP",
                                  font=("Arial", 11, "bold"),
                                  bg="#333333", fg="white",
                                  padx=16, pady=6, relief="flat",
                                  cursor="hand2", state="disabled",
                                  command=self.stop_recognition)
        self.stop_btn.pack(side="right", padx=5)

        self.start_btn = tk.Button(btn_frame, text="▶  START ATTENDANCE",
                                   font=("Arial", 11, "bold"),
                                   bg="#ff4757", fg="white",
                                   padx=16, pady=6, relief="flat",
                                   cursor="hand2",
                                   command=self.start_recognition)
        self.start_btn.pack(side="right", padx=5)

        self.count_label = tk.Label(bottom, text="Total Present: 0",
                                    font=("Arial", 12, "bold"),
                                    bg="#111111", fg="#2ed573")
        self.count_label.pack(side="right", padx=30, pady=10)

        # ---- MIDDLE CONTENT ----
        middle = tk.Frame(self.root, bg="#0d0d0d")
        middle.pack(fill="both", expand=True, padx=10, pady=8)

        # ---- LEFT: CAMERA (fixed size) ----
        left = tk.Frame(middle, bg="#111111")
        left.pack(side="left", fill="both", expand=True, padx=(0,8))

        tk.Label(left, text="LIVE CAMERA FEED",
                 font=("Arial", 9, "bold"),
                 bg="#111111", fg="#ff4757").pack(pady=(6,2))

        self.video_label = tk.Label(left, bg="#000000",
                                    width=640, height=480)
        self.video_label.pack(padx=8, pady=4)

        # ---- RIGHT: ATTENDANCE TABLE ----
        right = tk.Frame(middle, bg="#111111", width=370)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        tk.Label(right, text="ATTENDANCE LOG",
                 font=("Arial", 9, "bold"),
                 bg="#111111", fg="#ff4757").pack(pady=(8,2))

        today = datetime.now().strftime("%d %B %Y")
        tk.Label(right, text=f"{today}",
                 font=("Arial", 10),
                 bg="#111111", fg="#888888").pack(pady=(0,6))

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("F.Treeview",
                        background="#1a1a1a",
                        foreground="#ffffff",
                        fieldbackground="#1a1a1a",
                        font=("Arial", 10), rowheight=28)
        style.configure("F.Treeview.Heading",
                        background="#ff4757",
                        foreground="white",
                        font=("Arial", 10, "bold"))
        style.map("F.Treeview",
                  background=[("selected", "#ff4757")])

        columns = ("ID", "Name", "Roll", "Time", "Status")
        self.tree = ttk.Treeview(right, columns=columns,
                                 show="headings",
                                 style="F.Treeview")

        col_widths = {"ID":40, "Name":80, "Roll":75, "Time":70, "Status":75}
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=col_widths[col], anchor="center")

        self.tree.pack(padx=8, fill="both", expand=True, pady=6)

    def update_datetime(self):
        now = datetime.now().strftime("%d %b %Y   %H:%M:%S")
        self.datetime_label.config(text=now)
        self.root.after(1000, self.update_datetime)

    def start_recognition(self):
        self.running = True
        self.cap = cv2.VideoCapture(0)
        self.start_btn.config(state="disabled")
        self.stop_btn.config(state="normal")
        self.status_label.config(text="● Running — Detecting faces...",
                                  fg="#2ed573")
        attendance_marked.clear()
        self.update_frame()

    def stop_recognition(self):
        self.running = False
        if self.cap:
            self.cap.release()
        self.start_btn.config(state="normal")
        self.stop_btn.config(state="disabled")
        self.status_label.config(text="● Stopped", fg="#ff4757")
        self.video_label.config(image="", bg="#000000")

    def update_frame(self):
        if not self.running:
            return
        ret, frame = self.cap.read()
        if ret:
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, 1.3, 5)

            for (x, y, w, h) in faces:
                face = gray[y:y+h, x:x+w]
                face = cv2.resize(face, (100, 100))
                face_flat = face.flatten().reshape(1, -1)
                face_pca = pca.transform(face_flat)
                prediction = model.predict(face_pca)[0]
                confidence = model.predict_proba(face_pca).max()

                if confidence > 0.60:
                    details = student_details.get(prediction, {})
                    name = details.get("name", "Unknown")
                    roll = details.get("roll", "")
                    dept = details.get("dept", "")
                    color = (0, 255, 100)

                    cv2.rectangle(frame, (x, y), (x+w, y+h), color, 2)
                    cv2.putText(frame, f"ID:{prediction}",
                               (x, y-70), cv2.FONT_HERSHEY_SIMPLEX,
                               0.55, color, 2)
                    cv2.putText(frame, f"Name:{name}",
                               (x, y-50), cv2.FONT_HERSHEY_SIMPLEX,
                               0.55, color, 2)
                    cv2.putText(frame, f"Roll:{roll}",
                               (x, y-30), cv2.FONT_HERSHEY_SIMPLEX,
                               0.55, color, 2)
                    cv2.putText(frame, f"Dept:{dept}",
                               (x, y-10), cv2.FONT_HERSHEY_SIMPLEX,
                               0.55, color, 2)

                    marked = mark_attendance(prediction, name)
                    if marked:
                        now = datetime.now().strftime("%H:%M:%S")
                        self.tree.insert("", "end",
                                        values=(prediction, name,
                                                roll, now, "✓ Present"))
                        self.count_label.config(
                            text=f"Total Present: {len(attendance_marked)}")
                        self.status_label.config(
                            text=f"● Marked: {name}", fg="#2ed573")
                else:
                    cv2.rectangle(frame, (x, y), (x+w, y+h),
                                 (255, 71, 87), 2)
                    cv2.putText(frame, "Unknown", (x, y-10),
                               cv2.FONT_HERSHEY_SIMPLEX,
                               0.7, (255, 71, 87), 2)

            # Fixed camera size
            frame = cv2.resize(frame, (640, 460))
            frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            img = Image.fromarray(frame_rgb)
            imgtk = ImageTk.PhotoImage(image=img)
            self.video_label.imgtk = imgtk
            self.video_label.config(image=imgtk)

        self.root.after(10, self.update_frame)

    def clear_log(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        attendance_marked.clear()
        self.count_label.config(text="Total Present: 0")
        self.status_label.config(
            text="● Ready — Press START to begin", fg="#2ed573")

    def on_closing(self):
        self.stop_recognition()
        self.root.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = FaceTrackApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()